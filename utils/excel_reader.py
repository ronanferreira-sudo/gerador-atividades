"""
Utilitário para ler planilhas Excel e extrair dados das colunas:
- Cruzamento de: (Subfunção, Capacidade, Conhecimento)
- Identificação - MATRIZ DE REFERÊNCIA SAEP

O arquivo Excel deve ter o formato esperado de uma matriz de referência
onde a primeira linha ou primeiras linhas contém os cabeçalhos.
"""

import openpyxl
import re


def encontrar_linha_cabecalho(ws):
    """
    Varre as primeiras linhas da planilha para encontrar os cabeçalhos.
    Retorna o número da linha onde os cabeçalhos principais estão.
    """
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                valor = cell.value.strip().upper()
                if "MATRIZ DE REFERÊNCIA" in valor or "IDENTIFICAÇÃO" in valor:
                    return cell.row
    return 1


def encontrar_colunas(ws, linha_cabecalho):
    """
    Encontra as colunas relevantes na planilha:
    - Identificação - MATRIZ DE REFERÊNCIA SAEP
    - Cruzamento de: (Subfunção, Capacidade, Conhecimento)

    Retorna um dicionário com os índices das colunas encontradas.
    """
    colunas = {
        "identificacao": None,
        "subfuncao": None,
        "capacidade": None,
        "conhecimento": None,
        "cruzamento": None,  # Coluna que contém o cruzamento completo
    }

    # Procura nas linhas de cabeçalho (até 5 linhas a partir da linha do cabeçalho principal)
    for row in ws.iter_rows(
        min_row=max(1, linha_cabecalho - 5),
        max_row=linha_cabecalho + 5,
        max_col=ws.max_column,
    ):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                valor = cell.value.strip().upper()

                # Identificação - MATRIZ DE REFERÊNCIA SAEP
                if "IDENTIFICAÇÃO" in valor or "MATRIZ DE REFERÊNCIA" in valor:
                    if colunas["identificacao"] is None:
                        colunas["identificacao"] = cell.column

                # Cruzamento de: Subfunção
                if "SUBFUNÇÃO" in valor or "SUBFUNCAO" in valor:
                    colunas["subfuncao"] = cell.column

                # Cruzamento de: Capacidade
                if "CAPACIDADE" in valor:
                    colunas["capacidade"] = cell.column

                # Cruzamento de: Conhecimento
                if "CONHECIMENTO" in valor:
                    colunas["conhecimento"] = cell.column

                # Coluna de cruzamento (subfunção + capacidade + conhecimento juntos)
                if "CRUZAMENTO" in valor or ("SUB" in valor and "CAP" in valor and "CONHEC" in valor):
                    if colunas["cruzamento"] is None:
                        colunas["cruzamento"] = cell.column

    return colunas


def extrair_texto_celulas_mescladas(ws):
    """
    Para planilhas com células mescladas (formato PLANEJAMENTO DOCENTE),
    extrai o texto das células mescladas que contêm o conteúdo programático.
    """
    dados = []
    merged_ranges = list(ws.merged_cells.ranges)

    # Procura por linhas que contêm conteúdo na coluna A (ou colunas mescladas A:C)
    # Pula as primeiras linhas (cabeçalho) - começa da linha 13
    numero_aula = 1
    for row_idx in range(13, ws.max_row + 1):
        celula_a = ws.cell(row=row_idx, column=1)
        if celula_a.value and isinstance(celula_a.value, str) and len(celula_a.value.strip()) > 30:
            texto = celula_a.value.strip()

            # Ignora linhas que são cabeçalho (contém "Cruzamento de:" ou "O que?")
            texto_upper = texto.upper()
            if "CRUZAMENTO DE" in texto_upper or "O QUE?" in texto_upper:
                continue

            # Verifica se a célula está mesclada (formato PLANEJAMENTO DOCENTE)
            is_merged = any(
                celula_a.coordinate in merged_range
                for merged_range in merged_ranges
            )

            registro = {
                "identificacao": f"Aula {numero_aula}",
                "subfuncao": "",
                "capacidade": "",
                "conhecimento": texto,
                "tipo_planilha": "planejamento_docente",
            }
            dados.append(registro)
            numero_aula += 1

    return dados


def extrair_dados_excel(caminho_arquivo):
    """
    Lê o arquivo Excel e extrai os dados das colunas:
    - Identificação - MATRIZ DE REFERÊNCIA SAEP
    - Subfunção (dentro de Cruzamento de:)
    - Capacidade (dentro de Cruzamento de:)
    - Conhecimento (dentro de Cruzamento de:)

    Retorna uma lista de dicionários com os dados extraídos,
    apenas para linhas onde a coluna Identificação possui conteúdo.
    Também retorna um texto formatado para uso na IA.
    """
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb.active

    linha_cabecalho = encontrar_linha_cabecalho(ws)
    colunas = encontrar_colunas(ws, linha_cabecalho)

    print(f"📋 Linha do cabeçalho encontrada: {linha_cabecalho}")
    print(f"📋 Colunas encontradas: {colunas}")

    # Se nenhuma coluna foi encontrada, tenta abordagens alternativas
    if not any(colunas.values()):
        print("⚠️  Nenhuma coluna específica encontrada. Tentando formato PLANEJAMENTO DOCENTE...")
        dados = extrair_texto_celulas_mescladas(ws)
        if dados:
            wb.close()
            print(f"✅ Extraídos {len(dados)} itens no formato PLANEJAMENTO DOCENTE")
            return dados

        print("⚠️  Usando abordagem genérica...")
        dados = _extrair_generico(ws, linha_cabecalho)
        wb.close()
        return dados

    # Verifica se o conteúdo real está em células mescladas (ex: A13:C15)
    # Caso as colunas encontradas não tenham dados reais
    if colunas["subfuncao"] == colunas["capacidade"] == colunas["conhecimento"]:
        # Provavelmente é uma planilha com células mescladas onde tudo está na mesma coluna
        print("⚠️  Colunas de Subfunção, Capacidade e Conhecimento apontam para a mesma coluna.")
        print("⚠️  Tentando extrair como formato PLANEJAMENTO DOCENTE...")
        dados = extrair_texto_celulas_mescladas(ws)
        if dados:
            wb.close()
            print(f"✅ Extraídos {len(dados)} itens no formato PLANEJAMENTO DOCENTE")
            return dados

    dados = []
    linha_inicio = linha_cabecalho + 2  # Pular linha do cabeçalho e possivelmente uma linha em branco

    for row in ws.iter_rows(
        min_row=linha_inicio,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        # Pega o valor da coluna Identificação
        if colunas["identificacao"]:
            cell_id = row[colunas["identificacao"] - 1]
            valor_id = cell_id.value
        else:
            valor_id = None

        # Só considera linhas que tenham conteúdo na coluna Identificação
        if valor_id is not None and str(valor_id).strip():
            registro = {
                "identificacao": str(valor_id).strip(),
            }

            if colunas["subfuncao"]:
                cell_sub = row[colunas["subfuncao"] - 1]
                registro["subfuncao"] = str(cell_sub.value).strip() if cell_sub.value else ""

            if colunas["capacidade"]:
                cell_cap = row[colunas["capacidade"] - 1]
                registro["capacidade"] = str(cell_cap.value).strip() if cell_cap.value else ""

            if colunas["conhecimento"]:
                cell_conh = row[colunas["conhecimento"] - 1]
                registro["conhecimento"] = str(cell_conh.value).strip() if cell_conh.value else ""

            dados.append(registro)

    wb.close()
    return dados


def _extrair_generico(ws, linha_cabecalho):
    """
    Abordagem genérica: extrai todas as linhas com dados a partir do cabeçalho.
    Pega as primeiras colunas úteis.
    """
    dados = []
    linha_inicio = linha_cabecalho + 2

    for row in ws.iter_rows(
        min_row=linha_inicio,
        max_row=ws.max_row,
        min_col=1,
        max_col=min(ws.max_column, 4),  # Pega as primeiras 4 colunas
    ):
        valores = [cell.value for cell in row if cell.value is not None and str(cell.value).strip()]

        if valores:
            registro = {
                "identificacao": str(valores[0]).strip() if len(valores) > 0 else "",
                "subfuncao": str(valores[1]).strip() if len(valores) > 1 else "",
                "capacidade": str(valores[2]).strip() if len(valores) > 2 else "",
                "conhecimento": str(valores[3]).strip() if len(valores) > 3 else "",
            }
            if registro["identificacao"]:
                dados.append(registro)

    return dados


def formatar_dados_para_prompt(dados):
    """
    Formata os dados extraídos do Excel em um texto estruturado
    para ser usado como contexto no prompt da IA.
    Extrai os Conhecimentos, ou o texto completo quando não há separação.
    """
    if not dados:
        return "Nenhum dado encontrado no arquivo Excel."

    # Verifica se os dados são do formato PLANEJAMENTO DOCENTE
    if dados[0].get("tipo_planilha") == "planejamento_docente":
        linhas = []
        linhas.append("CONTEÚDO PROGRAMÁTICO DO CURSO")
        linhas.append("=" * 60)
        for i, registro in enumerate(dados, start=1):
            if registro.get("conhecimento") and registro["conhecimento"].strip():
                linhas.append(f"\n--- AULA {i} ---")
                linhas.append(registro["conhecimento"].strip())
        return "\n".join(linhas)

    conhecimentos = []
    for registro in dados:
        if registro.get("conhecimento") and registro["conhecimento"].strip():
            conhecimentos.append(registro["conhecimento"].strip())

    if not conhecimentos:
        # Se não encontrou conhecimentos separados, tenta juntar todo o texto disponível
        linhas = []
        linhas.append("CONTEÚDO DO CURSO")
        linhas.append("=" * 60)
        for i, registro in enumerate(dados, start=1):
            partes = []
            for campo in ["identificacao", "subfuncao", "capacidade", "conhecimento"]:
                if registro.get(campo) and registro[campo].strip():
                    partes.append(registro[campo].strip())
            if partes:
                linhas.append(f"\nItem {i}:")
                linhas.extend(partes)
        return "\n".join(linhas)

    linhas = []
    linhas.append("CONHECIMENTOS DO CURSO")
    linhas.append("=" * 60)

    for i, conhecimento in enumerate(conhecimentos, start=1):
        linhas.append(f"{i}. {conhecimento}")

    return "\n".join(linhas)


def contar_itens(dados):
    """Retorna a quantidade de Conhecimentos extraídos."""
    if not dados:
        return 0
    
    # Para formato PLANEJAMENTO DOCENTE, conta os itens com conhecimento
    if dados[0].get("tipo_planilha") == "planejamento_docente":
        return len([r for r in dados if r.get("conhecimento") and r["conhecimento"].strip()])
    
    conhecimentos = [r for r in dados if r.get("conhecimento") and r["conhecimento"].strip()]
    return len(conhecimentos)


if __name__ == "__main__":
    # Teste rápido
    import sys
    if len(sys.argv) > 1:
        caminho = sys.argv[1]
        dados = extrair_dados_excel(caminho)
        print(f"\n📊 Total de itens extraídos: {len(dados)}")
        print(f"📊 Total de conhecimentos: {contar_itens(dados)}")
        print("\n" + "=" * 60)
        print("TEXTO FORMATADO PARA IA:")
        print("=" * 60)
        print(formatar_dados_para_prompt(dados))